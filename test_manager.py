import pandas as pd
from openpyxl.workbook import Workbook

from tests import Tests
from excel_writer import ExcelWriter
import matplotlib
from ClusteringAnalyzer import ClusteringAnalyzer
import warnings
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from datetime import datetime
import os
from openpyxl.drawing.image import Image

warnings.filterwarnings("ignore")
matplotlib.use('Agg')


class TestManager:
    def __init__(self, df1, df2, output_folder, company_name, log_callback=None, P_A = 0.6):
        """
        Инициализация менеджера тестов.

        :param df1: DataFrame, содержащий данные из первого файла Excel
        :param df2: DataFrame, содержащий данные из второго файла Excel
        :param output_folder: Папка для сохранения результатов
        :param company_name: Имя компании для результатов
        :param log_callback: Функция для отправки логов в UI
        """
        self.df = df1
        self.OSV = df2.round(decimals=2).set_index('Счет')
        self.tests = Tests(self.df, self.OSV)
        self.excel_writer = ExcelWriter(output_folder, company_name)
        self.clustering_analyzer = ClusteringAnalyzer(self.df)
        self.log_callback = log_callback if log_callback else print  # Используем коллбек для логов или print по умолчанию
        self.filename = None
        self.output_folder = output_folder
        self.P_A = P_A

        current_date = datetime.now().strftime('%d-%m-%Y')
        # Если filename не передан, формируем его автоматически
        if not company_name:
            company_name = "default"
            self.filename = f"{company_name}_{current_date}.xlsx"
        else:
            self.filename = f"{company_name}_{current_date}.xlsx"
        
        self.file_path = os.path.join(self.output_folder, self.filename)
    
    def log(self, message):
        """Передаёт сообщение в лог."""
        self.log_callback(message)

    # Функция для анализа данных
    def analyze_data(self):
        from openpyxl import load_workbook
        import numpy as np
        import pandas as pd
        import matplotlib.pyplot as plt
        import tkinter as tk
        import os
        from tkinter import filedialog, messagebox

        """Кластеризация. Этап 1"""
        """

        СНАЧАЛА МЫ ПРОСТО РАССЧИТЫВАЕМ НЕКОТОРЫЕ ЗНАЧЕНИЯ ДЛЯ ПРОВЕДЕНИЯ КЛАСТЕРИЗАЦИИ
        И ПРОВОДИМ СТАНДАРТИЗАЦИЮ 


        """
        self.log("Подготовка к кластеризации...")
        # Добавляем признак: является ли сумма сторнированной (1) или нет (0)
        self.df["Сторно"] = self.df["Сумма"].apply(lambda x: 0 if x > 0 else 1)

        # Убираем строки с пустыми значениями поля Сумма
        self.df = self.df[self.df['Сумма'].notnull()]

        # Для отрицательных сумм меняем знак
        self.df.loc[self.df["Сумма"] < 0, "Сумма"] = -self.df["Сумма"]

        # Оставляем значения, больше 10, чтобы на них можно было провести все тесты
        self.df = self.df[self.df['Сумма'] > 10]

        # Меняем индексы
        self.df = self.df.reset_index(drop=True)

        temp = self.df.groupby('Сумма')['Сумма'].count() / len(self.df)
        sum_ = list(temp.index)
        frequency = list(temp)
        df_temp = pd.DataFrame({"Сумма": sum_, "Частота суммы": frequency})

        self.df = self.df.merge(df_temp, on='Сумма')
        test_overwrite = self.df

        self.df = test_overwrite
        print(self.df)
        import math
        from datetime import datetime
        # Вероятность появления цифры в первом разряде

        def p_first(d1):
            return math.log10(1 + 1 / d1)

        # Вероятность появления цифры во втором разряде

        def p_second(d2):
            s = 0
            for k in range(1, 10):
                s += math.log10(1 + 1 / (10 * k + d2))
            return s

        # Список, содержащий вероятности появления цифр 1-9 в первом #разряде

        first_teor = [p_first(d1) for d1 in range(1, 10)]
        df_duplicates = self.df.groupby(["СчетДт", "СчетКт", "Сумма"], as_index=False).count().sort_values(
            by="Организация",
            ascending=False)

        self.df.loc[:, 'first'] = self.df['Сумма'].apply(lambda x: int(str(x)[0]))

        # Список, содержащий частоты появления первых цифр в выборке

        first_real = self.df.groupby('first')['Сумма'].count() / len(self.df)

        # Расчет среднего абсолютного отклонения

        def MAD(AP, EP, k):
            s = 0
            for i in range(0, k - 1):
                s += abs(AP[i] - EP[i])
            return s / k

        mad = MAD(list(first_real), first_teor, 9)

        # Z-статистика

        def z_stat(AP, EP, N):
            chisl = abs(AP - EP)
            znam = ((EP * (1 - EP)) / N) ** 0.5
            if 1 / (2 * N) < chisl:
                chisl -= 1 / (2 * N)
            return chisl / znam

        # Z-Тест 1 цифры

        z_stats = []
        for i in range(9):
            z_stats.append(z_stat(list(first_real)[i], first_teor[i], len(self.df)))

        # Расчет хи-квадрат

        def chi2(AC, EC, N):
            k = len(AC)
            chi = 0
            for i in range(k):
                chi += (AC[i] * N - EC[i] * N) ** 2 / EC[i] * N
            return chi

        chi_stat = chi2(list(first_real), first_teor, len(self.df))

        # Добавление в исходный датафрейм столбца со значениями z-#статистик

        df_first_stats = pd.DataFrame({"first": list(range(1, 10)), "z-stat first": z_stats})
        self.df = self.df.merge(df_first_stats, on='first')

        # Z-Тест 2 цифры

        self.df.loc[:, 'second'] = self.df['Сумма'].apply(lambda x: int(str(x)[1]))

        second_real = self.df.groupby('second')['Сумма'].count() / len(self.df)

        second_teor = [p_second(d1) for d1 in range(0, 10)]

        z_stat_sec = []
        for i in range(10):
            z_stat_sec.append(z_stat(list(second_real)[i], second_teor[i], len(self.df)))

        df_second_stats = pd.DataFrame({"second": list(range(0, 10)), "z-stat second": z_stat_sec})
        self.df = self.df.merge(df_second_stats, on='second')

        # Тест первых двух цифр

        self.df.loc[:, 'first_two'] = self.df['Сумма'].apply(lambda x: int(str(x)[:2]))

        two_teor = [p_first(d1) for d1 in range(10, 100)]
        two_real = self.df.groupby('first_two')['Сумма'].count() / len(self.df)

        z_stat_two = []
        for i in range(90):
            z_stat_two.append(z_stat(list(two_real)[i], two_teor[i], len(self.df)))

        df_first_two_stats = pd.DataFrame({"first_two": list(range(10, 100)), "z-stat first_two": z_stat_two})
        self.df = self.df.merge(df_first_two_stats, on='first_two')

        # Тест суммирования

        two_real = self.df.groupby('first_two')['Сумма'].sum() / self.df['Сумма'].sum()

        df_abs_delta = pd.DataFrame({"first_two": list(range(10, 100)), "sum_frequency": list(two_real)})
        self.df = self.df.merge(df_abs_delta, on='first_two')

        # Тест второго порядка

        df_cur = self.df.sort_values(by='Сумма')
        df_cur.loc[:, 'two'] = df_cur['Сумма'].diff() * 10
        df_cur.dropna(subset=['Сумма'], inplace=True)
        df_cur = df_cur[df_cur['two'] > 10]
        df_cur.loc[:, 'two'] = df_cur['two'].apply(lambda x: int(str(x)[:2]))
        df_cur.shape

        df_z_stat_second_diff = pd.DataFrame({"two": list(range(10, 100)), "z_stat_second_diff": z_stat_two})

        df_cur.head()

        df_cur = df_cur.merge(df_z_stat_second_diff, on="two")

        ind = df_cur.index
        self.df.loc[ind, "z_stat_second_diff"] = df_cur["z_stat_second_diff"]

        # Z-Тест последних двух цифр

        df_cur = self.df

        df_cur.loc[:, 'last_two'] = df_cur['Сумма'].apply(lambda x: int(str(int(round((x * 100), 0)))[-2:]))

        two_real = df_cur.groupby('last_two')['Сумма'].count() / len(df_cur)

        two_teor = [0.01 for i in range(100)]

        z_stats = []
        for i in range(100):
            z_stats.append(z_stat(list(two_real)[i], two_teor[i], len(df_cur)))

        mad = MAD(list(two_real), two_teor, 100)

        df_last_two = pd.DataFrame({"last_two": list(range(0, 100)), "z_stat_last_two": z_stats})
        df_cur = df_cur.merge(df_last_two, on='last_two')

        def a_socr(a):
            return 10 * a / (10 ** int(math.log(a, 10)))

        df_cur['two'] = df_cur['Сумма'].apply(lambda x: a_socr(x))

        # Добавляется столбец с частотой счета Дт (синтетический счет)

        df_cur["СинтСчетДт"] = df_cur["СчетДт"].apply(lambda x: str(x)[:2])
        df_dt = (df_cur.groupby("СинтСчетДт").count() / len(df_cur))
        df_dt = df_dt.rename(columns={"Организация": "Частота счета Дт"})
        df_dt = df_dt["Частота счета Дт"]
        df_cur = df_cur.merge(df_dt, on='СинтСчетДт')

        # Добавляется столбец с частотой счета Кт (синтетический счет)

        df_cur["СинтСчетКт"] = df_cur["СчетКт"].apply(lambda x: str(x)[:2])
        df_kt = (df_cur.groupby("СинтСчетКт").count() / len(df_cur))
        df_kt = df_kt.rename(columns={"Организация": "Частота счета Кт"})
        df_kt = df_kt["Частота счета Кт"]
        df_cur = df_cur.merge(df_kt, on='СинтСчетКт')

        # Добавляется столбец с частотой проводки (по синтетическим счетам)

        df_cur["Проводка"] = df_cur["СинтСчетДт"] + "-" + df_cur["СинтСчетКт"]
        df_pr = (df_cur.groupby("Проводка").count() / len(df_cur))
        df_pr = df_pr.rename(columns={"Организация": "Частота проводки"})
        df_pr = df_pr["Частота проводки"]
        df_cur = df_cur.merge(df_pr, on="Проводка")

        # Добавляется столбец с частотой автора операции

        df_au = df_cur.groupby("АвторОперации").count() / len(df_cur)
        df_au = df_au.rename(columns={"Организация": "Частота автора операции"})
        df_au = df_au["Частота автора операции"]
        df_cur = df_cur.merge(df_au, on='АвторОперации')

        # Ручная проводка (1) или нет (0)

        df_cur["Ручная проводка"] = df_cur["РучнаяКорректировка"].apply(lambda x: 1 if x == "Да" else 0)

        # выходные дни (1), другие дни (0)

        df_cur["Data"] = df_cur["Период"].apply(lambda x: str(x)[:10])
        df_cur["Data"] = df_cur["Data"].apply(lambda x: datetime.strptime(x, "%Y-%m-%d"))
        df_cur["Data"] = df_cur["Data"].apply(lambda x: datetime.weekday(x))
        df_cur["Data"] = df_cur["Data"].apply(lambda x: 1 if (x == 5 or x == 6) else 0)
        df_cur = df_cur.rename(columns={"Data": "Выходные или рабочие"})

        # Считает количество дублирующихся проводок

        df_duplicates_ = df_duplicates.iloc[:, :4]
        df_duplicates_.rename({"Организация": "Количество дублей"}, axis=1, inplace=True)
        df_cur = df_cur.merge(df_duplicates_, on=["СчетДт", "СчетКт", "Сумма"])

        temp = df_cur.loc[:,
               ["z-stat first", "z-stat second", "z-stat first_two", "sum_frequency", "z_stat_second_diff",
                "z_stat_last_two",
                "Частота суммы", "Частота счета Дт", "Частота счета Кт", "Частота проводки",
                "Частота автора операции",
                "Ручная проводка", "Выходные или рабочие", "Сторно", "Количество дублей"]]

        temp.loc[temp["z_stat_second_diff"].isna(), "z_stat_second_diff"] = -1
        # Считает количество дублирующихся проводок
        df_duplicates_ = df_duplicates.iloc[:, :4]
        df_duplicates_.rename({"Организация": "Количество дублей"}, axis=1, inplace=True)
        df_cur = df_cur.merge(df_duplicates_, on=["СчетДт", "СчетКт", "Сумма"])

        temp.head()
        from sklearn import preprocessing
        scaled = preprocessing.StandardScaler().fit_transform(temp.iloc[:, :6].values)
        scaled_df = pd.DataFrame(scaled, index=temp.iloc[:, :6].index, columns=temp.iloc[:, :6].columns)
        scaled_df.head()
        plt.close('all')

        # Открываем существующий файл Excel
        workbook = load_workbook(self.file_path)

        # Имя нового листа
        new_sheet_name = 'Кластеризация'  # Укажите имя вашего нового листа

        # Проверяем, существует ли уже лист с таким названием
        if new_sheet_name not in workbook.sheetnames:
            workbook.create_sheet(new_sheet_name)  # Создаем новый пустой лист
        else:
            print(f'Лист с именем "{new_sheet_name}" уже существует.')

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        print(f'Пустой лист "{new_sheet_name}" успешно добавлен в файл {self.file_path}.')

        """

        ИЗ ЭТОГО ЭТАПА МЫ ДОЛЖНЫ ПОЛУЧИТЬ ИТОГОВЫЙ ДАТАФРЕЙМ scaled_df 
        ТАКЖЕ СОЗДАЁТСЯ ПУСТОЙ ЛИСТ ДЛЯ НАВИГАЦИИ ПО РАЗДЕЛАМ, ОН НУЖЕН

        НАЧИНАЕМ С ПОСТРОЕНИЯ СИЛУЭТА, ВЫВОДЫ: ТАБЛИЦА С КОЛИЧЕСТВОМ КЛАСТЕРОМ И Silhouette Score + ГРАФИК

        """

        import os
        import matplotlib.pyplot as plt
        from tqdm import tqdm
        from sklearn.cluster import KMeans
        import sklearn.metrics as metrics

        self.log("Построение силуэта...")

        # Настройки для графиков
        plt.rcParams["figure.figsize"] = (8, 4)

        # Список возможных значений кластеров
        x = [i for i in range(2, 6)]
        m = []

        sample_size = int(0.2 * scaled_df.shape[0])
        print(f'sample size = {sample_size}')

        # Подсчёт silhouette_score для каждого количества кластеров
        for i in x:
            labels = KMeans(n_clusters=i, random_state=1000).fit(scaled_df).labels_
            m.append(metrics.silhouette_score(scaled_df, labels, sample_size=sample_size))

        # Найдём количество кластеров с максимальным значением silhouette_score
        best_n_clusters = x[m.index(max(m))]

        print(f'Оптимальное количество кластеров: {best_n_clusters}')

        # Строим график зависимости silhouette_score от количества кластеров
        plt.plot(x, m, 'r-')
        plt.xticks(ticks=x, labels=[int(i) for i in x])
        plt.xlabel('Количество кластеров')
        plt.ylabel('Значение метрики')
        plt.title('Зависимость значения метрики от количества кластеров')

        # Сохраняем график
        image_path = 'silhouette_plot.png'
        plt.savefig(image_path)
        plt.close('all')  # Закрывает все открытые графики

        # Открываем существующий файл
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
        else:
            raise FileNotFoundError(f"Файл '{self.file_path}' не найден!")

        # Название нового листа для данных силуэта
        sheet_name = 'Этап 1'

        # Проверяем, существует ли уже лист "Силуэт", и создаём его, если нет
        if sheet_name not in workbook.sheetnames:
            worksheet_silhouette = workbook.create_sheet(sheet_name)
        else:
            worksheet_silhouette = workbook[sheet_name]

        # Записываем заголовки для данных
        worksheet_silhouette.cell(row=1, column=1, value='Количество кластеров')
        worksheet_silhouette.cell(row=1, column=2, value='Silhouette Score')

        # Записываем данные анализа с помощью метода append
        for n_clusters, score in zip(x, m):
            worksheet_silhouette.append([n_clusters, score])

        # Записываем строку с оптимальным числом кластеров
        worksheet_silhouette.cell(row=len(m) + 2, column=1,
                                  value=f'Оптимальное количество кластеров: {best_n_clusters}')

        from openpyxl.drawing.image import Image

        # Вставляем график в Excel
        image_path = 'silhouette_plot.png'
        img = Image(image_path)
        worksheet_silhouette.add_image(img, 'A14')

        # Автовыравнивание ширины столбцов A и B
        max_len_col1 = max([len(str(n)) for n in x] + [len('Количество кластеров')])
        max_len_col2 = max([len(f'{s:.3f}') for s in m] + [len('Silhouette Score')])

        worksheet_silhouette.column_dimensions['A'].width = max_len_col1
        worksheet_silhouette.column_dimensions['B'].width = max_len_col2

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временные файлы изображений
        if os.path.exists(image_path):
            os.remove(image_path)

        print(f"График силуэта и результаты анализа успешно добавлены в '{self.file_path}' на лист '{sheet_name}'.")

        """

        КЛАСТЕРИЗУЕМ ПО К-СРЕДНИХ, ВЫВОД: 
        1. ТАБЛИЦА "КЛАСС, ЧИСЛО ОБЪЕКТОВ"
        2. ТАБЛИЦА СРЕДНИХ ЗНАЧЕНИЙ Z-СТАТИСТИК
        3. ГРАФИК СРЕДНИХ ЗНАЧЕНИЙ


        """
        # Применение KMeans с оптимальным количеством кластеров
        if "Class" in scaled_df.columns:
            scaled_df.drop(columns=["Class"], inplace=True)

        km = KMeans(n_clusters=best_n_clusters, random_state=1000)

        scaled_df["Class"] = km.fit_predict(scaled_df)
        scaled_df["Class"] = scaled_df["Class"] + 1

        grouped_count = scaled_df.groupby("Class").count()["z-stat first"]
        print(grouped_count)

        temp["Class"] = scaled_df["Class"]

        # Выбираем только нужные столбцы для mean_temp
        selected_columns = ["z-stat first", "z-stat second", "z-stat first_two", "sum_frequency", "z_stat_second_diff",
                            "z_stat_last_two"]
        mean_temp = temp.groupby("Class")[selected_columns].mean()
        # Построение графика
        scaled_df.groupby("Class").mean().T.plot(grid=True, figsize=(15, 10),  # Увеличиваем высоту графика
                                                 rot=90,
                                                 xticks=range(len(scaled_df.columns) - 1),
                                                 style='o-', linewidth=4, markersize=12)

        plt.legend(fontsize=30)
        plt.tick_params(axis='x', labelsize=14)
        plt.tick_params(axis='y', labelsize=10)  # Уменьшаем шрифт оси y
        plt.tight_layout()  # Подбираем отступы

        # Сохраняем график
        image_path_z_stats = os.path.join(self.output_folder, 'z_stats.png')
        plt.savefig(image_path_z_stats)
        plt.close('all')  # Закрывает все открытые графики

        # Проверьте, что файл с графиком существует
        if not os.path.exists(image_path_z_stats):
            raise FileNotFoundError(f"Файл '{image_path_z_stats}' не был создан.")

        # Открываем существующий Excel-файл или создаём новый
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
        else:
            workbook = Workbook()

        # Название нового листа
        sheet_name = 'Этап 1'  # Изменено на 'Силуэт'

        # Проверяем, существует ли уже лист с названием "Силуэт"
        if sheet_name not in workbook.sheetnames:
            worksheet_silhouette = workbook.create_sheet(sheet_name)
        else:
            worksheet_silhouette = workbook[sheet_name]

        # Вставляем таблицу grouped_count в Excel
        worksheet_silhouette.cell(row=1, column=4, value='Класс')  # D1
        worksheet_silhouette.cell(row=1, column=5, value='Число объектов')  # E1

        # Записываем данные из grouped_count
        row_num = 2
        for class_label, count in grouped_count.items():
            worksheet_silhouette.cell(row=row_num, column=4, value=class_label)  # D
            worksheet_silhouette.cell(row=row_num, column=5, value=count)  # E
            row_num += 1

        # Вставляем заголовок для mean_temp
        worksheet_silhouette.cell(row=1, column=7, value='Средние значения')  # G1
        row_num = 2  # Начнем записи с G2 для классов

        # Записываем номера классов в G2, G3 и так далее
        for class_label in mean_temp.index:
            worksheet_silhouette.cell(row=row_num, column=6, value=class_label)  # Записываем номер класса в G
            row_num += 1

        # Записываем заголовки столбцов для mean_temp начиная с H1
        for col_num, col_name in enumerate(mean_temp.columns):
            worksheet_silhouette.cell(row=1, column=7 + col_num, value=col_name)  # Заголовки столбцов начинаются с H

        # Записываем значения для mean_temp
        row_num = 2  # Сбросим row_num для значений mean_temp
        for class_label, row in mean_temp.iterrows():
            for col_num, value in enumerate(row):
                worksheet_silhouette.cell(row=row_num, column=7 + col_num, value=value)  # Записываем значения
            row_num += 1

        from openpyxl.drawing.image import Image

        # Вставляем график в Excel
        img = Image(image_path_z_stats)
        img.width = img.width // 2.5  # Масштабирование изображения
        img.height = img.height // 2.5
        worksheet_silhouette.add_image(img, 'H14')

        # Устанавливаем автоподбор ширины для столбцов
        worksheet_silhouette.column_dimensions['D'].width = 15  # Ширина для столбца 'Класс'
        worksheet_silhouette.column_dimensions['E'].width = 20  # Ширина для столбца 'Число объектов'
        worksheet_silhouette.column_dimensions['G'].width = 20  # Ширина для столбца 'Средние значения'

        # Устанавливаем ширину для столбцов mean_temp начиная с H
        for col_num in range(len(mean_temp.columns)):
            worksheet_silhouette.column_dimensions[chr(72 + col_num)].width = 20  # Столбцы с заголовками начинаются с H

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временные файлы изображений
        if os.path.exists(image_path_z_stats):
            os.remove(image_path_z_stats)

        print(f"Лист 'Силуэт' обновлён в файле '{self.file_path}'. Файл готов для скачивания.")

        temp.groupby("Class").mean()

        """

        СТРОИМ ПРОСТО ГРАФИК ГЛАВНЫХ КОМПОНЕНТ И ВСЁ


        """

        self.log("Построение главных компонент...")

        from sklearn.decomposition import PCA
        pca = PCA(n_components=2)
        principalComponents = pca.fit_transform(scaled_df.iloc[:, :-1])
        principalDf = pd.DataFrame(data=principalComponents)

        principalDf = principalDf.rename(columns={0: "PC1", 1: "PC2"})
        principalDf["Class"] = scaled_df["Class"]

        principalDf.head()

        import pandas as pd
        import os
        import matplotlib.pyplot as plt
        from tqdm import tqdm
        from sklearn.cluster import KMeans
        import sklearn.metrics as metrics

        # Визуализация главных компонент
        for i in range(1, best_n_clusters + 1):
            data = principalDf[principalDf["Class"] == i]
            plt.plot(data.PC1, data.PC2, 'o', label=f'Класс {i}')
        plt.legend()
        plt.title("Главные компоненты")
        plt.xlabel("Главная компонента 1")
        plt.ylabel("Главная компонента 2")

        # Сохраняем график
        image_path_pca = os.path.join(self.output_folder, 'pca_plot.png')
        plt.savefig(image_path_pca)
        plt.close('all')  # Закрывает все открытые графики

        # Проверьте, что файл с графиком существует
        if not os.path.exists(image_path_pca):
            raise FileNotFoundError(f"Файл '{image_path_pca}' не был создан.")

        # Открываем существующий Excel-файл или создаём новый, если файла нет
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
        else:
            workbook = Workbook()

        # Название нового листа
        sheet_name = 'Этап 1'

        # Проверяем, существует ли уже лист с названием "Метод ГК"
        if sheet_name not in workbook.sheetnames:
            worksheet_pca = workbook.create_sheet(sheet_name)
        else:
            worksheet_pca = workbook[sheet_name]

        # Вставляем график PCA в Excel
        img = Image(image_path_pca)
        worksheet_pca.add_image(img, 'M14')

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временные файлы изображений
        if os.path.exists(image_path_pca):
            os.remove(image_path_pca)

        print(f"Лист 'Метод ГК' добавлен к файлу '{self.file_path}'. Файл готов для скачивания.")

        """

        АНАЛИЗ ВЫБРОСОВ, ВЫВОД:
        1. ТАБЛИЦЫ ("IsoLabels", "Ellabels")
        2. 2 ПОДПИСИ (Аномальный класс)
        3. ГРАФИК СРЕДНИХ ЗНАЧЕНИЙ ПО КЛАССАМ


        """

        self.log("Поиск аномального класса по Isolation Forest...")

        from sklearn.ensemble import IsolationForest
        anomaly_labels = IsolationForest().fit_predict(scaled_df.drop(["Class"], axis=1))

        scaled_df["IsoLabels"] = anomaly_labels

        # Список для хранения отношений z-stat first
        ratios = {}

        # Получаем уникальные классы
        classes = scaled_df["Class"].unique()

        # Проходим по каждому классу и вычисляем отношение
        for cls in classes:
            # Получаем count z-stat first для текущего класса
            z_stat_pos = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["IsoLabels"] == 1)][
                "z-stat first"].count()
            z_stat_neg = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["IsoLabels"] == -1)][
                "z-stat first"].count()

            # Проверяем, чтобы z_stat_neg не было нулем
            if z_stat_neg != 0:
                ratio = z_stat_pos / z_stat_neg
                ratios[cls] = ratio
            else:
                ratio = z_stat_pos
                ratios[cls] = ratio

        # Находим класс с минимальным отношением
        if ratios:
            anomaly_class_iso = min(ratios, key=ratios.get)  # Класс с минимальным отношением
        else:
            anomaly_class_iso = None  # Если нет классов

        print(f'Аномальный класс по Isolation Forest: {anomaly_class_iso}')

        self.log("Поиск аномального класса по Elliptic Envelope...")
        from sklearn.covariance import EllipticEnvelope
        anomaly_labels_el = EllipticEnvelope().fit_predict(scaled_df.drop(["Class", "IsoLabels"], axis=1))

        scaled_df["ElLabels"] = anomaly_labels_el
        scaled_df.groupby(["ElLabels", "IsoLabels", "Class"]).count()

        ratios_el = {}

        # Проходим по каждому классу и вычисляем отношение для Elliptic Envelope
        for cls in classes:
            # Получаем count z-stat first для текущего класса
            z_stat_pos = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["ElLabels"] == 1)][
                "z-stat first"].count()
            z_stat_neg = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["ElLabels"] == -1)][
                "z-stat first"].count()

            # Проверяем, чтобы z_stat_neg не было нулем
            if z_stat_neg != 0:
                ratio = z_stat_pos / z_stat_neg
                ratios_el[cls] = ratio
            else:
                ratio = z_stat_pos
                ratios_el[cls] = ratio

        # Находим аномальный класс для Elliptic Envelope
        if ratios_el:
            anomaly_class_el = min(ratios_el, key=ratios_el.get)  # Класс с минимальным отношением
        else:
            anomaly_class_el = None  # Если нет классов

        print(f'Аномальный класс по Elliptic Envelope: {anomaly_class_el}')

        import pandas as pd

        # Фильтрация аномальных операций по Isolation Forest
        anomalies_iso = scaled_df[(scaled_df["IsoLabels"] == -1) & (scaled_df["Class"] == anomaly_class_iso)]

        # Фильтрация аномальных операций по Elliptic Envelope
        anomalies_el = scaled_df[(scaled_df["ElLabels"] == -1) & (scaled_df["Class"] == anomaly_class_el)]

        # Объединение двух наборов данных по индексу (чтобы включить только общие строки)
        anomalies_combined = pd.merge(anomalies_iso, anomalies_el, how="inner", on=scaled_df.columns.tolist())

        # Группируем данные и заполняем нулями
        grouped_iso = scaled_df.groupby(["IsoLabels", "Class"]).count().reindex(
            pd.MultiIndex.from_product([[-1, 1], classes], names=["IsoLabels", "Class"]), fill_value=0
        )

        grouped_el = scaled_df.groupby(["ElLabels", "Class"]).count().reindex(
            pd.MultiIndex.from_product([[-1, 1], classes], names=["ElLabels", "Class"]), fill_value=0
        )

        # Открываем существующий Excel-файл или создаём новый, если файла нет
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
        else:
            workbook = Workbook()

        from openpyxl.utils import get_column_letter
        # Название нового листа
        sheet_name = 'Этап 1'  # Указываем лист "Силуэт"

        # Проверяем, существует ли уже лист с названием "Силуэт"
        if sheet_name not in workbook.sheetnames:
            worksheet_silhouette = workbook.create_sheet(sheet_name)
        else:
            worksheet_silhouette = workbook[sheet_name]

        # Записываем заголовки для Isolation Forest
        worksheet_silhouette.cell(row=1, column=14, value='IsoLabels')  # N1
        worksheet_silhouette.cell(row=1, column=15, value='Класс')  # O1
        worksheet_silhouette.cell(row=1, column=16, value='Число объектов')  # P1

        # Записываем данные для IsoLabels
        for row_num, (index, row) in enumerate(grouped_iso.iterrows(), start=2):
            worksheet_silhouette.cell(row=row_num, column=14, value=index[0])  # IsoLabels
            worksheet_silhouette.cell(row=row_num, column=15, value=index[1])  # Class
            worksheet_silhouette.cell(row=row_num, column=16, value=row['z-stat first'])  # Count

        # Добавляем информацию об аномальном классе по Isolation Forest
        anomaly_row = len(grouped_iso) + 2
        worksheet_silhouette.cell(row=anomaly_row, column=14, value='Аномальный класс по Isolation Forest:')
        worksheet_silhouette.cell(row=anomaly_row, column=15, value=anomaly_class_iso)

        # Записываем заголовки для Elliptic Envelope в Q1, R1 и S1
        worksheet_silhouette.cell(row=1, column=17, value='ElLabels')  # Q1
        worksheet_silhouette.cell(row=1, column=18, value='Класс')  # R1
        worksheet_silhouette.cell(row=1, column=19, value='Число объектов')  # S1

        # Записываем данные для ElLabels, начиная со строки 2
        for row_num, (index, row) in enumerate(grouped_el.iterrows(), start=2):
            worksheet_silhouette.cell(row=row_num, column=17, value=index[0])  # ElLabels
            worksheet_silhouette.cell(row=row_num, column=18, value=index[1])  # Class
            worksheet_silhouette.cell(row=row_num, column=19, value=row['z-stat first'])  # Count

        # Добавляем информацию об аномальном классе по Elliptic Envelope
        worksheet_silhouette.cell(row=len(grouped_el) + 2, column=17, value='Аномальный класс по Elliptic Envelope:')
        worksheet_silhouette.cell(row=len(grouped_el) + 2, column=18, value=anomaly_class_el)

        # Автовыравнивание ширины столбцов для IsoLabels и ElLabels
        for col in range(14, 20):  # N (14) до S (19)
            column_letter = get_column_letter(col)
            worksheet_silhouette.column_dimensions[column_letter].width = 20

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        print(f"Данные добавлены на лист 'Силуэт' в файле '{self.file_path}'.")

        scaled_df_temp = scaled_df.copy()
        scaled_df = scaled_df.drop(["ElLabels", "IsoLabels"], axis=1)

        from openpyxl.drawing.image import Image
        from openpyxl import load_workbook

        # Загружаем существующий файл Excel
        workbook = load_workbook(self.file_path)

        # Название нового листа для графиков
        sheet_name = 'Этап 1'

        # Проверяем, существует ли уже лист с таким названием
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
        else:
            worksheet = workbook[sheet_name]

        # Создание общего графика для всех классов и аномалий
        plt.figure(figsize=(15, 8))

        # Проходим по каждому классу и строим график среднего значения
        classes = scaled_df_temp["Class"].unique()
        for cls in classes:
            scaled_df_temp[scaled_df_temp["Class"] == cls].iloc[:, :-3].mean().T.plot(
                style='o-', linewidth=2, markersize=8, label=f'Среднее по классу {cls}'
            )

        # Добавляем график для аномалий (anomalies_combined)
        anomalies_combined.iloc[:, :-3].mean().T.plot(
            style='o--', linewidth=4, markersize=12, label='Среднее по аномальному классу'
        )

        # Настройки графика
        plt.legend(fontsize=16)  # Увеличиваем размер шрифта легенды
        plt.title('График средних значений по классам и аномальному классу', fontsize=20)  # Увеличиваем заголовок
        plt.xlabel('Показатели', fontsize=16)  # Увеличиваем размер шрифта для подписи по оси X
        plt.ylabel('Значение', fontsize=16)  # Увеличиваем размер шрифта для подписи по оси Y
        plt.xticks(fontsize=14)  # Увеличиваем размер шрифта для меток по оси X
        plt.yticks(fontsize=14)  # Увеличиваем размер шрифта для меток по оси Y
        plt.grid(True, linestyle='--', linewidth=0.5)

        # Сохранение графика в PNG
        image_file = 'combined_classes_plot.png'
        plt.savefig(image_file)
        plt.close('all')  # Закрывает все открытые графики

        # Вставляем изображение графика в ячейку листа Excel
        img = Image(image_file)
        img.width, img.height = img.width * 0.5, img.height * 0.5  # Уменьшаем масштаб графиков
        worksheet.add_image(img, 'S14')  # Вставляем график в ячейку

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временный файл с графиком после вставки в Excel
        os.remove(image_file)

        print(f'График успешно сохранён в файл {self.file_path}')

        """

        ИЗ ЭТОГО ЭТАПА МЫ ДОЛЖНЫ ПОЛУЧИТЬ ИТОГОВЫЙ ДАТАФРЕЙМ temp_class_1, ОН АНОМАЛЬНЫЙ


        """
        scaled_df = scaled_df_temp[(scaled_df_temp["ElLabels"] == -1) & (scaled_df_temp["IsoLabels"] == -1) & (
                scaled_df_temp["Class"] == anomaly_class_el) & (
                                           scaled_df_temp["Class"] == anomaly_class_iso)].iloc[:, :-2]

        cluster_1_index = list(scaled_df[(scaled_df["Class"] == anomaly_class_el) &
                                         (scaled_df["Class"] == anomaly_class_iso)].index)

        df_class_1 = temp[(temp.index.isin(cluster_1_index))]

        temp_class_1 = df_class_1.loc[:,
                       ["z-stat first", "z-stat second", "z-stat first_two", "sum_frequency", "z_stat_second_diff",
                        "Частота суммы", "Частота счета Дт", "Частота счета Кт", "Частота проводки",
                        "Частота автора операции",
                        "Ручная проводка", "Выходные или рабочие", "Сторно", "Количество дублей"]]

        """**3.2 Кластеризация, Этап 2**"""
        """

        ПОСТРОЕНИЕ СИЛУЭТА, ВЫВОДЫ: ТАБЛИЦА С КОЛИЧЕСТВОМ КЛАСТЕРОМ И Silhouette Score + ГРАФИК 


        """

        self.log("Построение силуэта аномального класса...")

        # стандартизация данных
        # проведем кластеризацию для объектов аномального класса по оставшимся признакам
        from sklearn.preprocessing import StandardScaler

        scaled_ = StandardScaler().fit_transform(temp_class_1.iloc[:, 6:].values)
        scaled_class_1 = pd.DataFrame(scaled_, index=temp_class_1.iloc[:, 6:].index,
                                      columns=temp_class_1.iloc[:, 6:].columns)
        scaled_class_1.head()

        # Силуэт
        plt.close('all')  # Закрывает все открытые графики

        plt.rcParams["figure.figsize"] = (8, 4)

        x = [i for i in range(2, 6)]
        m = []

        sample_size = int(scaled_class_1.shape[0])
        print(f'sample size = {sample_size}')
        for i in x:
            print(i)
            labels = KMeans(n_clusters=i, random_state=1000).fit(scaled_class_1).labels_
            print(labels)
            m.append(metrics.silhouette_score(scaled_class_1, labels, sample_size=sample_size))
            print(f'n = {i}, silhouette = {m[-1]}')
        # Найдём количество кластеров с максимальным значением silhouette_score
        best_n_clusters = x[m.index(max(m))]

        # Строим график зависимости silhouette_score от количества кластеров
        plt.plot(x, m, 'r-')
        plt.xticks(ticks=x, labels=[int(i) for i in x])
        plt.xlabel('Количество кластеров')
        plt.ylabel('Значение метрики')
        plt.title('Зависимость значения метрики от количества кластеров')

        # Сохраняем график
        image_path = 'silhouette_plot_2.png'
        plt.savefig(image_path)
        plt.close('all')  # Закрывает все открытые графики

        # Открываем существующий файл
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
        else:
            raise FileNotFoundError(f"Файл '{self.file_path}' не найден!")

        # Название нового листа для данных силуэта
        sheet_name = 'Этап 2'

        # Проверяем, существует ли уже лист "Силуэт", и создаём его, если нет
        if sheet_name not in workbook.sheetnames:
            worksheet_silhouette = workbook.create_sheet(sheet_name)
        else:
            worksheet_silhouette = workbook[sheet_name]

        # Записываем заголовки для данных
        worksheet_silhouette.cell(row=1, column=1, value='Количество кластеров')
        worksheet_silhouette.cell(row=1, column=2, value='Silhouette Score')

        # Записываем данные анализа с помощью метода append
        for n_clusters, score in zip(x, m):
            worksheet_silhouette.append([n_clusters, score])

        # Записываем строку с оптимальным числом кластеров
        worksheet_silhouette.cell(row=6, column=1, value=f'Оптимальное количество кластеров: {best_n_clusters}')

        img = Image(image_path)
        worksheet_silhouette.add_image(img, 'A14')

        # Автовыравнивание ширины столбцов A и B
        max_len_col1 = max([len(str(n)) for n in x] + [len('Количество кластеров')])
        max_len_col2 = max([len(f'{s:.3f}') for s in m] + [len('Silhouette Score')])

        worksheet_silhouette.column_dimensions['A'].width = max_len_col1
        worksheet_silhouette.column_dimensions['B'].width = max_len_col2

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временные файлы изображений
        if os.path.exists(image_path):
            os.remove(image_path)

        print(f"График силуэта и результаты анализа успешно добавлены в '{self.file_path}' на лист '{sheet_name}'.")
        plt.close('all')  # Закрывает все открытые графики

        """

        ПОЛУЧАЕМ СРЕДНИЕ ЗНАЧЕНИЯ ПО КЛАССАМ В ДАТАФРЕЙМ class_means + ГРАФИК


        """

        n_clusters = best_n_clusters
        if "Class" in scaled_class_1.columns:
            scaled_class_1.drop(columns=["Class", ], inplace=True)
        # Применение KMeans
        km = KMeans(n_clusters=n_clusters, random_state=1000)
        scaled_class_1["Class"] = km.fit_predict(scaled_class_1)
        scaled_class_1["Class"] = scaled_class_1["Class"] + 1  # Нумерация классов с 1

        # Вычисляем количество объектов в каждом классе
        class_counts = scaled_class_1.groupby("Class").count()["Сторно"]
        class_counts = class_counts.reset_index()  # Преобразуем индекс в столбец для удобного вывода
        class_counts.columns = ['Класс', 'Количество объектов']

        # Вычисляем средние значения по каждому классу
        class_means = scaled_class_1.groupby(
            "Class").mean().reset_index()  # Получаем средние значения и преобразуем в таблицу

        # Рисуем график средних значений по каждому классу
        scaled_class_1.groupby("Class").mean().T.plot(grid=True, figsize=(15, 10), rot=90,
                                                      xticks=range(len(scaled_class_1.columns) - 1), style='o-',
                                                      linewidth=4,
                                                      markersize=12)

        plt.legend(fontsize=30)
        plt.tick_params(axis='x', labelsize=18)
        plt.tick_params(axis='y', labelsize=18)

        # Сохраняем график во временный файл
        image_path = 'cluster_plot.png'
        plt.savefig(image_path, bbox_inches='tight')
        plt.close('all')  # Закрываем все открытые графики

        # Название нового листа для кластеризации
        sheet_name = 'Этап 2'

        # Открываем существующий файл Excel
        workbook = load_workbook(self.file_path)

        # Проверяем, существует ли уже лист с таким названием
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
        else:
            worksheet = workbook[sheet_name]

        # Добавляем заголовки для данных количества объектов, начиная с ячейки D1
        worksheet.cell(row=1, column=4, value="Класс")  # D1
        worksheet.cell(row=1, column=5, value="Число объектов")  # E1

        # Записываем данные по количеству объектов в каждом классе в Excel
        for idx, row in class_counts.iterrows():
            worksheet.append(row.tolist())  # Записываем каждую строку данных

        # Добавляем заголовки для средних значений начиная с колонки D
        start_col = 4  # Колонка D
        worksheet.cell(row=1, column=start_col, value="Класс")
        for i, col_name in enumerate(class_means.columns[1:], start=start_col + 1):
            worksheet.cell(row=1, column=i, value=col_name)

        # Записываем средние значения в таблицу, начиная со строки 2
        for idx, row in class_means.iterrows():
            for col_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=idx + 2, column=col_idx, value=value)

        # Вставляем график в ячейку M7
        img = Image(image_path)
        img.width, img.height = img.width * 0.4, img.height * 0.4  # Уменьшаем масштаб графика
        worksheet.add_image(img, 'M14')  # Вставляем график в ячейку M7

        # Сохраняем изменения в файл Excel
        workbook.save(self.file_path)

        # Удаляем временный файл с графиком
        if os.path.exists(image_path):
            os.remove(image_path)

        print(f'График и данные кластеров успешно добавлены в файл {self.file_path}.')

        """

        ВЫВОДИМПОДОЗРИТЕЛЬНЫЕ ОПЕРАЦИИ ПУТЁМ РАЗДЕЛЕНИЯ ДАТАФРЕЙМА НА КУСОЧКИ И 
        ПООЧЕРЁДНОГО СОХРАНЕНИЯ КУСОЧКОВ В ЭКСЕЛЬ, ПОТОМ РАСКРАШИВАЕМ ЯЧЕЙКИ


        """

        from openpyxl.styles import Font, PatternFill

        # Цвета для каждого класса
        class_colors = {
            "Класс 1": "0000FF",  # Синий
            "Класс 2": "FFA500",  # Оранжевый
            "Класс 3": "008000",  # Зеленый
            "Класс 4": "FF0000",  # Красный
            "Класс 5": "800080"  # Фиолетовый
        }

        sheet_name = "Подозрительные операции"
        output_data = []  # Для хранения данных перед записью в Excel

        # Записываем данные по каждому кластеру
        for class_num in sorted(scaled_class_1["Class"].unique()):
            # Добавляем заголовок класса
            output_data.append([f"Класс {class_num}"])  # Заголовок

            # Извлекаем строки для данного класса
            class_df = self.df.loc[scaled_class_1[scaled_class_1["Class"] == class_num].index]

            # Добавляем строки данного кластера в список
            for idx, row in class_df.iterrows():
                output_data.append(row.tolist())  # Каждая строка в новый ряд

        # Преобразуем данные в DataFrame для удобства записи
        output_df = pd.DataFrame(output_data)

        # Сохраняем данные в существующий Excel-файл на новый лист
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Записываем данные на новый лист
            output_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        # Загружаем workbook, чтобы применить форматирование
        workbook = load_workbook(self.file_path)
        worksheet = workbook[sheet_name]

        # Применяем форматирование к заголовкам "Класс X"
        bold_font = Font(size=14, bold=True)
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            is_class_row = False  # Переменная для отслеживания, является ли строка "Класс"
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("Класс"):
                    is_class_row = True  # Устанавливаем флаг, если "Класс" найден
                    break  # Выходим из внутреннего цикла

            if is_class_row:  # Если строка содержит "Класс"
                for cell in row:  # Применяем к каждой ячейке в строке
                    cell.font = bold_font
                    # Устанавливаем цвет фона, если он задан для данного класса
                    color_code = class_colors.get(cell.value)
                    if color_code:
                        cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

        # Сохраняем изменения в файл
        workbook.save(self.file_path)
        print(f"Данные с разделением по классам успешно записаны в лист '{sheet_name}' файла '{self.file_path}'.")

        """

        ЭТО КУСОК СТАРОГО ИНТЕРФЕЙСА, ОН ТЕБЕ НЕ ПРИГОДИТСЯ 100%


        """

        # Функция для выбора вероятностей в зависимости от значений MAD
        def get_mad_probabilities(mad_value, test_type):
            # MAD для первой цифры
            if test_type == "first_digit":
                if 0.000 <= mad_value <= 0.006:
                    return 0, 0.99
                elif 0.006 < mad_value <= 0.012:
                    return 0.33, 0.66
                elif 0.012 < mad_value <= 0.015:
                    return 0.66, 0.33
                else:
                    return 0.99, 0

            # MAD для второй цифры
            elif test_type == "second_digit":
                if 0.000 <= mad_value <= 0.008:
                    return 0, 0.99
                elif 0.008 < mad_value <= 0.010:
                    return 0.33, 0.66
                elif 0.010 < mad_value <= 0.012:
                    return 0.66, 0.33
                else:
                    return 0.99, 0

            # MAD для первых двух цифр
            elif test_type == "first_two_digits":
                if 0.0000 <= mad_value <= 0.0012:
                    return 0, 0.99
                elif 0.0012 < mad_value <= 0.0018:
                    return 0.33, 0.66
                elif 0.0018 < mad_value <= 0.0022:
                    return 0.66, 0.33
                else:
                    return 0.99, 0

            # MAD второго порядка
            elif test_type == "second_order":
                if 0.0000 <= mad_value <= 0.0012:
                    return 0, 0.99
                elif 0.0012 < mad_value <= 0.0018:
                    return 0.33, 0.66
                elif 0.0018 < mad_value <= 0.0022:
                    return 0.66, 0.33
                else:
                    return 0.99, 0

            # Если тип теста не известен
            else:
                raise ValueError("Неизвестный тип теста.")

        workbook = load_workbook(self.file_path)

        # Функция для извлечения значения MAD из строки "Среднее абсолютное отклонение (MAD): *значение*"
        def get_mad_value(sheet, cell):
            mad_cell_value = sheet[cell].value
            if mad_cell_value and "MAD" in mad_cell_value:
                mad_value = mad_cell_value.split(":")[-1].strip()  # Извлечение значения после двоеточия
                return float(mad_value)  # Преобразование в число
            return None

        # Извлечение значений MAD для каждого теста
        mad_first_digit = get_mad_value(workbook['Тест 2.1'], 'A5')
        mad_second_digit = get_mad_value(workbook['Тест 2.2'], 'A5')
        mad_first_two_digits = get_mad_value(workbook['Тест 2.3'], 'A5')
        mad_second_order = get_mad_value(workbook['Тест 2.3'], 'A5')
        if mad_second_order is not None:
            mad_second_order = float(mad_second_order)
        # Вывод значений MAD
        print(f"MAD для первой цифры: {mad_first_digit}")
        print(f"MAD для второй цифры: {mad_second_digit}")
        print(f"MAD для первых двух цифр: {mad_first_two_digits}")
        print(f"MAD для второго порядка: {mad_second_order}")

        # Функция для расчета вероятности фрода
        def bayesian_fraud_risk():
            # Ввод данных от пользователя
            P_B_A_first, P_B_not_A_first = get_mad_probabilities(mad_first_digit, "first_digit")
            P_B_A_second, P_B_not_A_second = get_mad_probabilities(mad_second_digit, "second_digit")
            P_B_A_first_two, P_B_not_A_first_two = get_mad_probabilities(mad_first_two_digits, "first_two_digits")
            P_B_A_second_order, P_B_not_A_second_order = get_mad_probabilities(mad_second_order, "second_order")

            # Рассчитаем средние вероятности для всех тестов
            P_B_A = (P_B_A_first + P_B_A_second + P_B_A_first_two + P_B_A_second_order) / 4
            P_B_not_A = (P_B_not_A_first + P_B_not_A_second + P_B_not_A_first_two + P_B_not_A_second_order) / 4

            # Вычисляем P(B)
            P_B = P_B_A * self.P_A + P_B_not_A * (1 - self.P_A)

            # Проверка на деление на ноль
            if P_B == 0:
                print("Ошибка: P(B) равно нулю, невозможно вычислить P(A|B).")
                return None, None

            # Вычисляем вероятность фрода при данных тестах (P(A|B))
            P_A_B = (P_B_A * self.P_A) / P_B
            print(f'Вероятность фрода при данных тестах: P(A|B) = {P_A_B:.4f}')

            P_A_B *= 100  # Переводим в проценты

            # Классификация риска
            if 0 <= P_A_B < 25:
                return P_A_B, "низкий"
            elif 25 <= P_A_B < 50:
                return P_A_B, "пониженный"
            elif 50 <= P_A_B < 75:
                return P_A_B, "повышенный"
            elif 75 <= P_A_B <= 100:
                return P_A_B, "высокий"
            else:
                return P_A_B, "Что-то пошло не так"

        # Запуск программы и получение значений P_A_B и risk_level
        p_a_b_value, risk_level = bayesian_fraud_risk()  # Сохраняем уровень риска для записи в файл

        # Создаем новый лист "Оценка риска"
        risk_sheet = workbook.create_sheet(title="Оценка риска")

        # Запись данных в новый лист с изменённым шрифтом
        risk_sheet['A1'] = f"Вероятность фрода при вероятности P(A) = {self.P_A} имеет значение: {p_a_b_value:.4f}%."
        risk_sheet['A1'].font = bold_font  # Применение шрифта к A1

        risk_sheet['A2'] = f"Уровень риска: {risk_level}."
        risk_sheet['A2'].font = bold_font  # Применение шрифта к A2

        # Сохраняем изменения в файле Excel
        workbook.save(self.file_path)

        print(f"Данные успешно добавлены на лист 'Оценка риска' в файле {self.file_path}.")

    def run_tests(self):
        """
        Запускает все тесты и записывает результаты в Excel.
        """
        self.log("Запуск тестов согласованности данных...")
        self.excel_writer.save_data_to_excel(pd.DataFrame(), 'Тесты целостности')
        correct_data = True

        # Выполнение тестов
        consistency_result, is_success = self.tests.test_coherence_data()
        self.excel_writer.save_data_to_excel(consistency_result, 'Тест 1.1')
        if not is_success:
            correct_data = False

        self.log("Запуск тестов математической корректности...")
        math_correct, is_success = self.tests.test_math_correctly()
        self.excel_writer.save_data_to_excel(math_correct, 'Тест 1.2')
        if not is_success:
            correct_data = False

        self.log("Запуск тестов полноты выгрузки...")
        unloading_complete, is_success = self.tests.test_unloading_completeness()
        self.excel_writer.save_data_to_excel(unloading_complete, 'Тест 1.3')
        if not is_success:
            correct_data = False

        if correct_data is False:
            raise Exception('файлы с данными некорректны. Подробная информация в Excel отчете')

        self.log("Запуск теста Бенфорда...")
        benford_result = self.tests.benford_check()
        self.excel_writer.save_data_to_excel(benford_result, 'Закон Бенфорда')

        i = 1
        self.log("Запуск теста первой, второй и первой и второй цифры...")
        digit_results = self.tests.test_digits()
        for name, df in digit_results.items():
            plot_file = f'{name}_plot.png'
            self.excel_writer.save_data_with_charts(df, f'Тест 2.{i}', plot_file)
            i = i + 1

        self.log("Запуск теста суммирования...")
        df_results, image_path = self.tests.test_summation()
        self.excel_writer.save_data_with_charts(df_results, 'Тест 2.4', image_path)

        self.log("Запуск теста второго порядка...")
        df_results, image_path = self.tests.test_sec_order()
        self.excel_writer.save_data_with_charts(df_results, 'Тест 2.5', image_path)

        self.log("Запуск теста мантисс...")
        df_results, image_paths = self.tests.test_mantiss()
        self.excel_writer.save_data_with_charts(df_results, 'Тест 2.6', image_paths)

        self.log("Запуск теста дублирования сумм...")
        df_results, image_path = self.tests.test_amount_duplication()
        self.excel_writer.save_data_with_charts(df_results, 'Тест 2.7', image_path)

        self.log("Запуск теста двух последних цифр...")
        df_results, image_path = self.tests.test_two_last_digit()
        self.excel_writer.save_data_with_charts(df_results, 'Тест 2.8', image_path)

        self.log("Запуск расчета коэффициента искажения...")
        df_results = self.tests.calculate_coef_distortion()
        self.excel_writer.save_data_to_excel(df_results, 'Тест 2.9')

        # self.log("Запуск вычисления силуэта")
        # df_results, image_path = self.clustering_analyzer.calculate_digit_stats()
        # self.excel_writer.save_data_with_charts(df_results, 'Силуэт', image_path)
        #
        # self.log("Запуск вычисления статистик")
        # df_results, image_path = self.clustering_analyzer.calculate_statistics()
        # self.excel_writer.save_data_with_charts(df_results, 'Статистики', image_path)
        #
        # self.log("Построение главных компонент...")
        # df_results, image_path = self.clustering_analyzer.calculate_principal_components()
        # self.excel_writer.save_data_with_charts(df_results, 'Метод ГК', image_path)
        #
        # self.log("Поиск аномального класса по Isolation Forest..")
        # df_results = self.clustering_analyzer.calculate_anomaly_class()
        # self.excel_writer.save_data_with_charts(df_results, 'Аномальный класс (этап 1)')
        #
        # self.log("Построение графиков")
        # df_results, image_paths = self.clustering_analyzer.create_graphics()
        # self.excel_writer.save_data_with_charts(df_results, 'Графики', image_paths)
        #
        # self.log("Поиск подозрительных операций")
        # df_results = self.clustering_analyzer.search_suspicious_transactions_1()
        # self.excel_writer.save_data_with_charts(df_results, 'Подозрительные операции 2')
        #
        # self.log("Расчет силуэта (этап 2)")
        # df_results, image_path = self.clustering_analyzer.calculate_digit_stats_2()
        # self.excel_writer.save_data_with_charts(df_results, 'Силуэт (этап 2)', image_path)
        #
        # self.log("Кластеризация (этап 2)")
        # df_results, image_path = self.clustering_analyzer.calculate_clasterization_2()
        # self.excel_writer.save_data_with_charts(df_results, 'Кластеризация (этап 2)', image_path)
        #
        # self.log("Поиск аномального класса (этап 2)")
        # df_results = self.clustering_analyzer.search_anomaly_classes()
        # self.excel_writer.save_data_with_charts(df_results, 'Аномальный класс (этап 2)')
        #
        # self.log("Вывод подозрительных операций (этап 2)")
        # df_results = self.clustering_analyzer.search_suspicious_transactions_2()
        # self.excel_writer.save_data_with_charts (df_results, 'Подозрительные операции 2')

        self.analyze_data()

        self.excel_writer.delete_png_files()
        self.log("Все тесты завершены.")






